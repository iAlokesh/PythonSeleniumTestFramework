import openpyxl


class TestDataReader:

    @staticmethod
    def get_testdata(tc_name):
        Dict = {}
        wb = openpyxl.load_workbook("C:\\Users\\aloke\\Desktop\\data.xlsx")
        sheet = wb.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == tc_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        print(Dict)
        return Dict
